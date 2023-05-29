import pandas as pd 
from openpyxl import load_workbook 
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import json
import string
# import discord
import logging
# from discord import SyncWebhook


class ExcelReportPlugin():
    def __init__(self,
                 input_file,
                 output_file):
        self.input_file = input_file
        self.output_file = output_file
    
    def main(self):
        df = self.read_input_file()
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime("%Y-%m-%d")
        df_transform = self.transform(df)
        self.create_output_file(df_transform)
        print('Workbook Created!')

        # SELECT ACTIVE WORKSHEET 
        wb = load_workbook(self.output_file)
        wb.active = wb['Report']

        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row

        self.column_dimension(wb.active)
        self.barchart (wb.active, min_row, max_row, max_column, min_column)
        self.add_total(wb.active, max_row, min_row, max_column)
        self.save_file(wb)

    def read_input_file(self):
        df = pd.read_excel(self.input_file)
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime("%Y-%m-%d")
        print(df.head())
        return df
    
    def transform(self, df:pd.DataFrame) -> pd.DataFrame:
        df_transform = df.pivot_table(index=['Gender','Date'],
        columns='Product line', 
        values='Total', 
        aggfunc='sum').round()
        print(df_transform)
    
    def create_output_file(self, df_transform):
        df_transform.to_excel(self.output_file, 
        sheet_name='Report', 
        startrow=4)

        print('Save dataframe done...')

    def barchart(self, workbook, min_column, max_column, min_row, max_row):
        barchart = BarChart() 

        data = Reference(workbook.active,
                        min_col=min_column+1,
                        max_col=max_column,
                        min_row=min_row,
                        max_row=max_row
                        )

        categories = Reference(workbook.active,
                            min_col=min_column,
                            max_col=max_column,
                            min_row=min_row+1,
                            max_row=max_row
                            )

        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categories)

        workbook.active.add_chart(barchart, 'J5')
        barchart.title = 'Sales berdasarkan Prodk'
        barchart.width = 50
        barchart.height = 15
        barchart.style = 2
    
    def column_dimension(self, worbook):
        dim_holder = DimensionHolder(worksheet=worbook.active)

    def add_total(self, workbook, max_row, min_row, max_column):
        alphabet = list(string.ascii_uppercase) # [1, 2, 3, dst ....  ]
        alphabet_excel = alphabet[:max_column]  # [A, B, C, D, E, F, G]
        for i in alphabet_excel:
            if i != 'A':
                workbook[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
                workbook[f'{i}{max_row+1}'].style = 'Currency'

        workbook[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

        workbook['A1'] = 'Sales Report'
        workbook['A2'] = '2019'
        workbook['A1'].font = Font('Arial', bold=True, size=20)
        workbook['A2'].font = Font('Arial', bold=True, size=10)
        
        def save_file(self, wb):
            wb.save(self.output_file)
            print('File Saved!')




# class ExcelReportPlugin():
    def __init__(self,
                 input_file,
                 output_file
                 ):
        self.input_file = input_file
        self.output_file = output_file

    def main(self):
        df = self.read_input_file()
        df_transform = self.transform(df)
        self.create_output_file(df_transform)
        print("workbook created")

        wb = load_workbook(self.output_file)
        wb.active = wb['Report']

        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row
        
        self.column_dimension(wb.active)
        self.barchart(wb.active, min_column, max_column, min_row, max_row)
        self.add_total(max_column, max_row, min_row, wb.active)
        self.save_file(wb)

    def read_input_file(self):
        df = pd.read_excel(self.input_file)
        logging.info(df.head())
        return df


    def transform(self, df:pd.DataFrame) -> pd.DataFrame:
        df_transform = df.pivot_table(index='Gender', 
                                    columns='Product line', 
                                    values='Total', 
                                    aggfunc='sum').round()
        return df_transform


    def create_output_file(self, df):
        print('Save dataframe to excel...')
        df.to_excel(self.output_file, 
                        sheet_name='Report', 
                        startrow=4)
        print(f'Save dataframe done... {self.output_file}')

    def column_dimension(self, workbook):
        dim_holder = DimensionHolder(worksheet=workbook)

        for col in range(workbook.min_column, workbook.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(workbook, min=col, max=col, width=20)

        workbook.column_dimensions = dim_holder


    def barchart(self, workbook, min_column, max_column, min_row, max_row):
        barchart = BarChart()

        data = Reference(workbook, 
                        min_col=min_column+1,
                        max_col=max_column,
                        min_row=min_row,
                        max_row=max_row
                        )

        categories = Reference(workbook,
                                min_col=min_column,
                                max_col=min_column,
                                min_row=min_row+1,
                                max_row=max_row
                                )

        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categories)


        workbook.add_chart(barchart, 'B12')
        barchart.title = 'Sales berdasarkan Produk'
        barchart.style = 2


    def add_total(self, max_column, max_row, min_row, wb):
        alphabet = list(string.ascii_uppercase)
        alphabet_excel = alphabet[:max_column]
        #[A,B,C,D,E,F,G]
        for i in alphabet_excel:
            if i != 'A':
                wb[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
                wb[f'{i}{max_row+1}'].style = 'Currency'

        wb[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

        wb['A1'] = 'Sales Report'
        wb['A2'] = '2019'
        wb['A1'].font = Font('Arial', bold=True, size=20)
        wb['A2'].font = Font('Arial', bold=True, size=10)

    def save_file(self, wb):
        wb.save(self.output_file)
        print('File saved')
